import sqlite3
import openpyxl as xl
import os
from datetime import datetime

con = sqlite3.connect("game.db")

cur = con.cursor()

cur.execute("""
    CREATE TABLE IF NOT EXISTS game(   
        game_id INTEGER PRIMARY KEY,
        name,
        system,
        notes
    )
"""
)

def insert_game(name, system, notes):
    cur.execute("""
        INSERT INTO game(name, system, notes) VALUES(
            "{}",
            "{}",
            "{}"
        )
    """.format(name, system, notes)
    )
    
    con.commit()
    
def show_results(query):
    results = cur.execute(query)
    games = results.fetchall()
    
    if len(games) == 0:
        print("Could not find games starting with that letter")
    else:
        for game in games:
            print("{:<4} | {:<77} \n{}".format(game[0], game[1], game[3]))
            print("=" * 80)

def get_all_games():
    query = """
        SELECT game_id, name, system, notes
        FROM game
    """
    show_results(query)

def get_games_for_system(system):
    query = """
        SELECT game_id, name, system, notes
        FROM game
        WHERE system = '{}'
    """.format(system)
    show_results(query)

def get_games_for_system_letter(system, letter):
    query = """
        SELECT game_id, name, system, notes
        FROM game
        WHERE system = '{}'
        AND name like '{}%' COLLATE NOCASE
        ORDER BY name COLLATE NOCASE
    """.format(system, letter)
    show_results(query)
    
def get_games_for_system_search(search):
    query = """
        SELECT game_id, name, system, notes
        FROM game
        WHERE name like '%{}%' COLLATE NOCASE
        ORDER BY name COLLATE NOCASE
    """.format(search)
    show_results(query)
    
def get_game_by_id(game_id):
    query = """
        SELECT game_id, name, system, notes
        FROM game
        WHERE game_id = {}
    """.format(game_id)
    show_results(query)
    
def update_game_name(game_name, game_id):
    cur.execute("""
        UPDATE game 
        SET name = '{}'
        WHERE game_id = {}
    """.format(game_name, game_id)
    )
    con.commit()

def update_game_notes(game_notes, game_id):
    cur.execute("""
        UPDATE game 
        SET notes = '{}'
        WHERE game_id = {}
    """.format(game_notes, game_id)
    )
    con.commit()
    
def get_systems(): 
    wb = xl.load_workbook(filename="game_data.xlsx")
    system_names = wb.sheetnames
    wb.close()
    systems = [""] + system_names
    systems_to_delete = ["PriceTotals", "Systems", "Accessories", "Wish List", "Kelly", "TODO", "Blue_TODO", "Loss"]
    systems = [system for system in systems if system not in systems_to_delete]
    return systems

def migrate_system_to_db(system):
    wb = xl.load_workbook(filename="game_data.xlsx")
    sheet = wb[system]
    
    for row_header in sheet.iter_rows(values_only=True, min_row=1, max_row=1):
        lower_row = [item.lower() for item in row_header if item is not None]
        name_index = lower_row.index("name")
        notes_index = lower_row.index("notes")
        
    game_count = 0
    for row in sheet.iter_rows(values_only=True, min_row=2):
        if row[name_index] is not None:
            game_name = row[name_index]
            notes = ""
            if row[notes_index] is not None:
                notes = row[notes_index]
                
            print(game_name, " -> ", notes)
            
            game_notes = notes.replace('"', '_')
            insert_game(game_name, system, game_notes)
            game_count = game_count + 1

    print("count:", game_count )
    wb.close()
    
def show_headers():
    wb = xl.load_workbook(filename="game_data.xlsx")
    systems = get_systems()
    for system in systems:
        if system == "" : 
            continue
        sheet = wb[system]
        
        print(system)
        for row in sheet.iter_rows(values_only=True, min_row=1, max_row=1):
            lower_row = [item.lower() for item in row if item is not None]
            print(lower_row)
            print(lower_row.index("name"), lower_row.index("notes"))

    wb.close()

def load_initial_data():
    systems = get_systems()
    for system in systems:
        if system == "":
            continue
        print("System:", system)
        migrate_system_to_db(system)

def backup_database():
    if not os.path.exists("database_backups"):
        os.mkdir("database_backups")
    
    timestamp = datetime.now().strftime('%m_%d_%Y_%I_%M_%p')
    backup_file = "database_backups\{}_{}.sql".format("game_database", timestamp)
    print("Backing up database to file:\n\t", backup_file)
    with open(backup_file, 'w') as f:
        for line in con.iterdump():
            f.write('%s\n' % line)
    